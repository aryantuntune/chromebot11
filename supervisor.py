"""Supervisor: runs run_bot.py repeatedly until every target account has an OTP.

Usage (from chromebot11 directory):
    .venv\Scripts\python.exe supervisor.py

What it does:
  1. Launches run_bot.py as a subprocess.
  2. When the bot exits, re-reads OTP_results.xlsx to count captured vs retryable.
  3. If retryable accounts remain AND the last run made progress -> restart after a
     short cooldown (BOT_BETWEEN_RUNS_SEC, default 60s).
  4. If the last run made NO progress (portal throttling / all stuck) -> waits a
     longer cooldown (BOT_STALL_COOLDOWN_SEC, default 300s) before restarting.
  5. After BOT_MAX_STALLS (default 3) consecutive no-progress runs it gives up and
     reports which accounts are still stuck.
  6. Writes a status file (output/supervisor_status.json) after every run so the
     dashboard can display supervisor-level stats.

Environment overrides:
    BOT_INPUT_FILE         path to the input .xlsx (default: auto-detect from input/)
    BOT_BETWEEN_RUNS_SEC   cooldown between runs when progress was made (default: 60)
    BOT_STALL_COOLDOWN_SEC cooldown when a run made zero progress (default: 300)
    BOT_MAX_STALLS         how many consecutive no-progress runs before giving up (default: 3)
    BOT_MAX_PASSES         max retry passes PER bot run (default: 6, passed to run_bot.py)
"""
from __future__ import annotations

import json
import logging
import os
import pathlib
import re
import subprocess
import sys
import time

from openpyxl import load_workbook

# ── Single-instance lock ──────────────────────────────────────────────────────
_LOCK_FILE = pathlib.Path("output/supervisor.lock")

def _acquire_lock() -> bool:
    """Write our PID to the lock file. Return False if another instance is running."""
    _LOCK_FILE.parent.mkdir(parents=True, exist_ok=True)
    if _LOCK_FILE.exists():
        try:
            other_pid = int(_LOCK_FILE.read_text().strip())
            # Check if that PID is still alive
            import ctypes
            handle = ctypes.windll.kernel32.OpenProcess(0x0400, False, other_pid)
            if handle:
                ctypes.windll.kernel32.CloseHandle(handle)
                return False   # still running
        except Exception:
            pass  # stale lock — take it
    _LOCK_FILE.write_text(str(os.getpid()))
    return True

def _release_lock() -> None:
    try:
        _LOCK_FILE.unlink(missing_ok=True)
    except Exception:
        pass

# ── Config ──────────────────────────────────────────────────────────────────
INPUT_FILE          = os.environ.get("BOT_INPUT_FILE", "").strip()
BETWEEN_RUNS_SEC    = int(os.environ.get("BOT_BETWEEN_RUNS_SEC",    "60"))
STALL_COOLDOWN_SEC  = int(os.environ.get("BOT_STALL_COOLDOWN_SEC", "300"))
MAX_STALLS          = int(os.environ.get("BOT_MAX_STALLS",           "3"))
MAX_PASSES_PER_RUN  = int(os.environ.get("BOT_MAX_PASSES",           "6"))

OUTPUT_DIR   = pathlib.Path("output")
OTP_RESULTS  = OUTPUT_DIR / "OTP_results.xlsx"
STATUS_FILE  = OUTPUT_DIR / "supervisor_status.json"

_RETRY_MARKERS = (
    "NO_BOOKING_TABLE", "NO_BOOKINGS", "EMPTY_STATUS",
    "CAPTURE_ERROR", "ERROR", "CAPTCHA_FAILED",
)

PYTHON = str(pathlib.Path(sys.executable))


# ── Helpers ──────────────────────────────────────────────────────────────────

def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [SUPERVISOR] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _find_input_file() -> pathlib.Path:
    """Auto-detect the most recently modified master xlsx in input/."""
    if INPUT_FILE and pathlib.Path(INPUT_FILE).exists():
        return pathlib.Path(INPUT_FILE)
    candidates = [
        p for p in pathlib.Path("input").glob("*.xlsx")
        if not p.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError("No .xlsx files in input/")
    masters = []
    for p in candidates:
        try:
            wb = load_workbook(p, read_only=True)
            cols = wb.active.max_column or 1
            wb.close()
            if cols >= 3:
                masters.append(p)
        except Exception:
            pass
    if not masters:
        raise FileNotFoundError("No master credentials file found in input/")
    return max(masters, key=lambda p: p.stat().st_mtime)


def _load_target_ids(path: pathlib.Path) -> list[str]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    ids = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        v = row[0]
        if v not in (None, ""):
            ids.append(str(v).strip())
    wb.close()
    return ids


def _read_results(target_ids: set[str]) -> dict[str, str]:
    """Return {consumer_id: result} for every target ID found in OTP_results.xlsx."""
    if not OTP_RESULTS.exists():
        return {}
    try:
        wb = load_workbook(str(OTP_RESULTS), read_only=True, data_only=True)
        ws = wb.active
        results: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            cid = str(row[0] or "").strip()
            val = str(row[1] or "").strip()
            if cid in target_ids:
                results[cid] = val
        wb.close()
        return results
    except Exception:
        return {}


def _count(results: dict[str, str], target_ids: list[str]):
    captured, retryable, other, missing = 0, 0, 0, 0
    for cid in target_ids:
        val = results.get(cid, "")
        if not val:
            missing += 1
        elif re.fullmatch(r"\d{3,8}", val):
            captured += 1
        elif any(val.upper().startswith(m) for m in _RETRY_MARKERS):
            retryable += 1
        else:
            other += 1   # Delivered, In process, NOT_IN_MASTER, etc.
    return captured, retryable, other, missing


def _write_status(data: dict) -> None:
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        STATUS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")
    except Exception:
        pass


def _stall_cooldown(secs: int, log) -> None:
    log.warning("No progress last run — cooling down %ds before retry...", secs)
    for remaining in range(secs, 0, -10):
        log.info("  Cooldown: %ds remaining...", remaining)
        time.sleep(min(10, remaining))


def _short_cooldown(secs: int, log) -> None:
    log.info("Short cooldown %ds before next run...", secs)
    time.sleep(secs)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    _setup_logging()
    log = logging.getLogger("supervisor")

    if not _acquire_lock():
        log.error("Another supervisor is already running (see output/supervisor.lock). Exiting.")
        return 1
    import atexit
    atexit.register(_release_lock)

    input_path = _find_input_file()
    log.info("Input file  : %s", input_path.name)

    target_ids  = _load_target_ids(input_path)
    target_set  = set(target_ids)
    total       = len(target_ids)
    log.info("Targets     : %d accounts", total)

    run_no       = 0
    stall_streak = 0
    prev_captured = -1

    while True:
        run_no += 1

        # ── Snapshot BEFORE this run ──────────────────────────────────────
        results_before = _read_results(target_set)
        cap_before, ret_before, oth_before, miss_before = _count(results_before, target_ids)
        still_todo = ret_before + miss_before

        log.info("=== Supervisor run %d ===  captured=%d  retryable=%d  missing=%d",
                 run_no, cap_before, ret_before, miss_before)

        if still_todo == 0:
            log.info("All accounts captured or finalized. Done!")
            break

        _write_status({
            "run": run_no, "state": "running",
            "total": total, "captured": cap_before,
            "retryable": ret_before, "missing": miss_before,
            "stall_streak": stall_streak,
        })

        # ── Launch bot subprocess ─────────────────────────────────────────
        env = os.environ.copy()
        env["BOT_MAX_PASSES"] = str(MAX_PASSES_PER_RUN)
        env["BOT_START_ROW"]  = "2"
        env.setdefault("BOT_BROWSER",     "edge")
        env.setdefault("BOT_SETTLE_MS",   "5000")
        env.setdefault("BOT_BETWEEN_MS",  "5000")
        env.setdefault("BOT_AUTO_RETRY",  "1")
        env.setdefault("BOT_ADAPTIVE",    "1")

        log.info("Launching run_bot.py  (MAX_PASSES=%d)...", MAX_PASSES_PER_RUN)
        try:
            proc = subprocess.run(
                [PYTHON, "run_bot.py"],
                env=env,
            )
        except KeyboardInterrupt:
            log.warning("Interrupted by user. Saving current status.")
            break

        # ── Snapshot AFTER this run ───────────────────────────────────────
        results_after  = _read_results(target_set)
        cap_after, ret_after, oth_after, miss_after = _count(results_after, target_ids)
        newly_captured = cap_after - cap_before
        still_todo_now = ret_after + miss_after

        log.info("Run %d done: +%d captured this run  (total captured=%d, remaining=%d)",
                 run_no, newly_captured, cap_after, still_todo_now)

        _write_status({
            "run": run_no, "state": "idle",
            "total": total, "captured": cap_after,
            "retryable": ret_after, "missing": miss_after,
            "newly_captured": newly_captured,
            "stall_streak": stall_streak,
        })

        if still_todo_now == 0:
            log.info("All %d accounts captured. Supervisor done!", total)
            break

        # ── Progress check ────────────────────────────────────────────────
        if newly_captured == 0:
            stall_streak += 1
            log.warning("No progress this run (stall %d/%d).", stall_streak, MAX_STALLS)
            if stall_streak >= MAX_STALLS:
                # List the stuck accounts
                stuck = [
                    (cid, results_after.get(cid, "NOT_YET_RUN"))
                    for cid in target_ids
                    if (results_after.get(cid, "") == ""
                        or any(results_after.get(cid, "").upper().startswith(m)
                               for m in _RETRY_MARKERS))
                ]
                log.error(
                    "Giving up after %d consecutive no-progress runs. "
                    "%d account(s) still stuck:", MAX_STALLS, len(stuck)
                )
                for cid, val in stuck:
                    log.error("  Consumer %s -> %r", cid, val)
                _write_status({
                    "run": run_no, "state": "stuck",
                    "total": total, "captured": cap_after,
                    "retryable": ret_after, "missing": miss_after,
                    "stuck": stuck,
                })
                break
            _stall_cooldown(STALL_COOLDOWN_SEC, log)
        else:
            stall_streak = 0
            if still_todo_now > 0:
                _short_cooldown(BETWEEN_RUNS_SEC, log)

    # ── Final summary ─────────────────────────────────────────────────────
    results_final = _read_results(target_set)
    cap, ret, oth, miss = _count(results_final, target_ids)
    log.info("=== FINAL SUMMARY ===  total=%d  captured=%d  other=%d  retryable=%d  missing=%d",
             total, cap, oth, ret, miss)
    if cap == total - oth:
        log.info("SUCCESS — all processable accounts have an OTP.")
    else:
        log.warning("%d account(s) did not produce an OTP.", ret + miss)

    return 0


if __name__ == "__main__":
    sys.exit(main())
