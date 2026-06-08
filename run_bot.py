"""Replay the recorded HP Gas login flow for every row of the input Excel.

This runner reproduces, per spreadsheet row, the exact sequence of actions
captured with the Playwright recorder (see ``recorded_flow.py``):

  1. Open the HP Gas portal login page.
  2. Fill the Mobile/E-Mail field with the row's email (column B / ``id``).
  3. PAUSE: the human solves the CAPTCHA in the visible Chrome window and
     presses <Enter> in the CAPTCHA box (this reveals the password field),
     then returns to THIS terminal and presses <Enter> to continue.
  4. Fill the password (column C) once the password field appears.
  5. Click Login, dismiss the "OK" confirmation dialog (best-effort), and
     click "View Cylinder Booking history" (best-effort).
  6. Record a result string into the ``code`` column (column D) and save a
     copy of the workbook into ``output/`` after every row.

Run it from the project root, in YOUR OWN terminal (so the Enter-pause works):

    .venv/bin/python run_bot.py

Each row opens a fresh, logged-out browser context. Press Ctrl-C any time to
stop; progress is saved incrementally after each completed row.
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import pathlib
import re
import sys
import time

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError

# --------------------------------------------------------------------------- #
# Settings (kept inline; the YAML config model does not fit this staged flow).
# --------------------------------------------------------------------------- #
LOGIN_URL = "https://myhpgas.in/myHPGas/PortalLogin.aspx"
# Headed by default. With auto-CAPTCHA no human is needed, so you can run
# invisibly with BOT_HEADLESS=1.
HEADLESS = os.environ.get("BOT_HEADLESS", "").strip().lower() in ("1", "true", "yes")

# Which browser to drive. Override with BOT_BROWSER=chrome|edge|brave|opera.
# The booking page on this portal is flaky; if one browser loads it poorly,
# try another. All are Chromium-based, so the automation is identical.
BROWSER = os.environ.get("BOT_BROWSER", "chrome").strip().lower()

NAV_TIMEOUT_MS = 90_000
ELEMENT_TIMEOUT_MS = 45_000
# The booking-history page is server-rendered and can be slow; give the grid
# plenty of time to appear after we click into it before we read the OTP.
BOOKING_NAV_TIMEOUT_MS = 120_000

# How long to wait for the human to solve the CAPTCHA (the password field
# appearing is the signal that the CAPTCHA was accepted).
CAPTCHA_SOLVE_TIMEOUT_MS = 300_000  # 5 minutes per row

# Selectors captured by the recorder.
SEL_EMAIL = "#ContentPlaceHolder1_txtUserNameEmail"
SEL_CAPTCHA = "#ContentPlaceHolder1_loginCaptcha_tbCaptchaInput"
SEL_CAPTCHA_IMG = "#ContentPlaceHolder1_loginCaptcha_imgCaptcha"
SEL_CAPTCHA_REFRESH = "#ContentPlaceHolder1_loginCaptcha_btnRefreshCaptcha"
SEL_PASSWORD = "#ContentPlaceHolder1_txtPassword"
SEL_LOGIN_BTN = "#ContentPlaceHolder1_btnLogin"

# Booking-history grid shown after login. Newest booking is the FIRST data row.
# 0-based <td> indices in a row:
#   0=Order Date 1=Order Ref No 2=Order No 3=No.of Cyl 4=Status 5=Date
#   6=Cash Memo No 7=Cash Memo Amount 8=Cash Memo Date 9=delivery/OTP cell ...
# The delivery cell (index 9) shows "OutForDelivery (OTP: 7149)" while the
# cylinder is out for delivery, or "Delivered" once delivered. The Status cell
# (index 4) shows "In process" / "Delivered". The OTP we want is in cell 9.
SEL_BOOKING_TABLE = "#ContentPlaceHolder1_gvBookingHistory"
BOOKING_ROW = "tr.GridviewScrollItem"
STATUS_COL_IDX = 4       # 0-based <td> index of the Status cell
ORDERREF_COL_IDX = 1     # 0-based <td> index of the Order Ref No cell
DELIVERY_COL_IDX = 9     # 0-based <td> index of the "OutForDelivery (OTP: …)" cell

# --------------------------------------------------------------------------- #
# File layout. Two input workbooks live in input/:
#   * MASTER  -- the big SHREELALJI credentials DB (>=5 columns, no header):
#         A=conno(consumer id)  B=name  C=email(login id)  D=Hpgas-pwd  E=password
#   * TARGETS -- a single column of consumer IDs to actually process this run.
# We match each target consumer ID against column A of the master to pull that
# account's login email (C) + password (E).
# --------------------------------------------------------------------------- #
INPUT_DIR = "input"
OUTPUT_DIR = "output"

MASTER_SHEET = "Sheet1"
MASTER_CONNO_COL = "A"     # consumer id (matched against the targets)
MASTER_EMAIL_COL = "C"     # login id (email / mobile)
MASTER_PWD_COL = "D"       # login password

TARGETS_CONNO_COL = "A"    # the targets file's consumer-id column

OTP_RESULTS_NAME = "OTP_results.xlsx"   # separate output: Consumer ID | OTP

# Optional run controls via environment variables (run.bat stays simple):
#   BOT_MAX_ROWS=5  -> only process the first 5 target IDs (great for a test run)
#   BOT_DEBUG=1     -> dump screenshot + page text/html per account to output/debug/
MAX_ROWS = int(os.environ["BOT_MAX_ROWS"]) if os.environ.get("BOT_MAX_ROWS") else None
DEBUG_DUMP = os.environ.get("BOT_DEBUG", "").strip() not in ("", "0", "false", "no")

# Human-like pacing (milliseconds). The portal can choke on rapid automated
# steps and needs a moment AFTER login to finish setting up the session before
# the heavy booking-history page will render. These pauses mimic a person's
# pace and cut down on the booking grid failing to load. Tune via env vars.
SETTLE_MS = int(os.environ.get("BOT_SETTLE_MS", "3000"))    # after login, before opening booking
BETWEEN_MS = int(os.environ.get("BOT_BETWEEN_MS", "3000"))  # pause between accounts

# Auto-solve the CAPTCHA with OCR (ddddocr) instead of a human. The portal
# allows unlimited tries, so a wrong guess just refreshes and we try again.
AUTO_CAPTCHA = os.environ.get("BOT_AUTO_CAPTCHA", "1").strip().lower() not in ("0", "false", "no")
CAPTCHA_MAX_TRIES = int(os.environ.get("BOT_CAPTCHA_TRIES", "15"))
# Save every confirmed-correct (image, text) pair, building a training set that
# the solver can be fine-tuned on over time.
COLLECT_CAPTCHA = os.environ.get("BOT_CAPTCHA_DATASET", "1").strip().lower() not in ("0", "false", "no")

# Parallel sharding: split the targets so each worker/browser takes a unique,
# contiguous slice -- no consumer ID is ever processed by two workers, and no
# login is ever reused across browsers. BOT_OUT_SUFFIX keeps per-shard output
# files from clashing; BOT_MASTER_INDEX lets a shard load a prebuilt id->creds
# map (built once by the parallel launcher) instead of re-loading the big master.
SHARD_INDEX = int(os.environ.get("BOT_SHARD_INDEX", "0"))         # 0-based
SHARD_COUNT = max(1, int(os.environ.get("BOT_SHARD_COUNT", "1")))
OUT_SUFFIX = os.environ.get("BOT_OUT_SUFFIX", "")
MASTER_INDEX = os.environ.get("BOT_MASTER_INDEX", "").strip()

# Adaptive pacing: automatically ride the portal's throttle. The delay between
# logins GROWS after booking-page failures (likely throttling) and SHRINKS after
# a run of clean successes; a sustained failure streak triggers a cooldown so
# the server can recover. Set BOT_ADAPTIVE=0 for a fixed BOT_BETWEEN_MS instead.
ADAPTIVE = os.environ.get("BOT_ADAPTIVE", "1").strip().lower() not in ("0", "false", "no")
MIN_BETWEEN_MS = int(os.environ.get("BOT_MIN_BETWEEN_MS", "3000"))
MAX_BETWEEN_MS = int(os.environ.get("BOT_MAX_BETWEEN_MS", "30000"))
COOLDOWN_FAILS = int(os.environ.get("BOT_COOLDOWN_FAILS", "5"))    # consecutive fails -> cooldown
COOLDOWN_MS = int(os.environ.get("BOT_COOLDOWN_MS", "120000"))     # cooldown length (ms)

# Auto-retry: after a full pass, keep re-passing (skip captured, retry failures)
# until everything's captured or a pass makes no progress, up to BOT_MAX_PASSES.
AUTO_RETRY = os.environ.get("BOT_AUTO_RETRY", "1").strip().lower() not in ("0", "false", "no")
MAX_PASSES = int(os.environ.get("BOT_MAX_PASSES", "4"))

# Self-contained single-file mode: a workbook that carries its own IDs, emails
# and passwords (header row). BOT_START_ROW / BOT_END_ROW restrict to a 1-based,
# inclusive spreadsheet row range so you can resume / cap a partially-done file.
START_ROW = int(os.environ.get("BOT_START_ROW", "0"))
END_ROW = int(os.environ.get("BOT_END_ROW", "0"))
# Force a specific self-contained file (used by the overnight orchestrator).
FILE_OVERRIDE = os.environ.get("BOT_FILE", "").strip()
# Overnight resilience: on a pass that captures nothing, cool down and retry
# instead of stopping; only give up after this many consecutive dry passes.
DRY_STOP = int(os.environ.get("BOT_DRY_STOP", "3"))


def _setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _wait_for_captcha_solved(page, email: str) -> None:
    """Block until the human solves the CAPTCHA in the browser.

    The signal is the password field becoming visible: on this site the
    password box only appears AFTER the CAPTCHA is typed and accepted (Enter in
    the CAPTCHA box). We poll for it so there is no dependency on terminal input
    — the operator interacts only with the visible Chrome window.
    """
    banner = (
        "\n"
        "============================================================\n"
        f"  ACTION REQUIRED  —  signing in as: {email}\n"
        "------------------------------------------------------------\n"
        "  In the visible Chrome window:\n"
        "    1. Type the CAPTCHA code shown on the page.\n"
        "    2. Press <Enter> in the CAPTCHA box.\n"
        "  The PASSWORD field then appears and the bot continues\n"
        "  automatically — nothing to do in this terminal.\n"
        "============================================================\n"
    )
    sys.stdout.write(banner)
    sys.stdout.flush()

    deadline = time.monotonic() + (CAPTCHA_SOLVE_TIMEOUT_MS / 1000.0)
    pwd = page.locator(SEL_PASSWORD).first
    while time.monotonic() < deadline:
        try:
            if pwd.is_visible():
                return
        except Exception:  # noqa: BLE001 - transient DOM states during reload
            pass
        time.sleep(0.5)
    raise TimeoutError(
        "CAPTCHA not solved in time: the password field "
        f"({SEL_PASSWORD!r}) never appeared within "
        f"{CAPTCHA_SOLVE_TIMEOUT_MS // 1000}s."
    )


# --------------------------------------------------------------------------- #
# Automatic CAPTCHA solving (OCR) with self-collected training data.
# --------------------------------------------------------------------------- #
_OCR = None
_CAP_STATS = {"attempts": 0, "successes": 0, "first_try": 0, "accounts_solved": 0}


def _get_ocr():
    """Lazily build the ddddocr solver once (imported here so manual mode needs
    no extra dependency)."""
    global _OCR
    if _OCR is None:
        import ddddocr
        # Use a locally fine-tuned model + charset if train_captcha.py has
        # produced one; otherwise the bundled model. Output is post-filtered to
        # [a-z0-9] either way (this portal's CAPTCHA alphabet).
        model = pathlib.Path(OUTPUT_DIR) / "captcha_model"
        onnx, charsets = model / "model.onnx", model / "charsets.json"
        if onnx.exists() and charsets.exists():
            _OCR = ddddocr.DdddOcr(show_ad=False, import_onnx_path=str(onnx),
                                   charsets_path=str(charsets))
            logging.getLogger("bot").info("CAPTCHA: using fine-tuned model %s", onnx)
        else:
            _OCR = ddddocr.DdddOcr(show_ad=False)
    return _OCR


def _refresh_captcha(page) -> None:
    """Click the CAPTCHA refresh link to load a fresh image."""
    try:
        page.locator(SEL_CAPTCHA_REFRESH).click()
        page.wait_for_timeout(700)
    except Exception:  # noqa: BLE001 - best-effort
        pass


def _save_captcha_sample(img_bytes: bytes, label: str) -> None:
    """Save a CONFIRMED-correct CAPTCHA as ``<label>_<md5>.png`` for training.

    Success proves the label is correct, so this builds a free, accurately
    labelled dataset that a custom model can later be fine-tuned on.
    """
    if not COLLECT_CAPTCHA:
        return
    try:
        d = pathlib.Path(OUTPUT_DIR) / "captcha_dataset"
        d.mkdir(parents=True, exist_ok=True)
        h = hashlib.md5(img_bytes).hexdigest()[:10]
        (d / f"{label}_{h}.png").write_bytes(img_bytes)
    except Exception:  # noqa: BLE001 - best-effort
        pass


def _save_captcha_stats() -> None:
    """Write rolling solver accuracy so improvement over time is visible."""
    try:
        s = dict(_CAP_STATS)
        s["overall_solve_rate_pct"] = round(
            (s["successes"] / s["attempts"] * 100) if s["attempts"] else 0, 1)
        s["first_try_rate_pct"] = round(
            (s["first_try"] / s["accounts_solved"] * 100) if s["accounts_solved"] else 0, 1)
        (pathlib.Path(OUTPUT_DIR) / "captcha_stats.json").write_text(
            json.dumps(s, indent=2))
    except Exception:  # noqa: BLE001 - best-effort
        pass


def _auto_solve_captcha(page) -> bool:
    """Read and submit the CAPTCHA with OCR until the password field appears.

    The portal allows unlimited tries: a wrong guess keeps the same image, so we
    click refresh for a new one and retry, up to ``CAPTCHA_MAX_TRIES``. Each
    confirmed-correct image is saved for training and stats are updated.
    """
    log = logging.getLogger("bot")
    ocr = _get_ocr()
    pwd = page.locator(SEL_PASSWORD).first
    img_loc = page.locator(SEL_CAPTCHA_IMG)
    for attempt in range(1, CAPTCHA_MAX_TRIES + 1):
        try:
            img = img_loc.screenshot()
        except Exception:  # noqa: BLE001 - page not ready
            return False
        guess = re.sub(r"[^a-z0-9]", "", ocr.classification(img).lower())
        _CAP_STATS["attempts"] += 1
        if len(guess) != 6:
            # The CAPTCHA is always 6 chars; a different length means a misread.
            _refresh_captcha(page)
            continue
        try:
            page.fill(SEL_CAPTCHA, guess)
            page.locator(SEL_CAPTCHA).press("Enter")
            pwd.wait_for(state="visible", timeout=4_000)
        except PlaywrightTimeoutError:
            _refresh_captcha(page)   # wrong guess -> fresh image, try again
            continue
        except Exception:  # noqa: BLE001 - transient; refresh and retry
            _refresh_captcha(page)
            continue
        _CAP_STATS["successes"] += 1
        _CAP_STATS["accounts_solved"] += 1
        if attempt == 1:
            _CAP_STATS["first_try"] += 1
        _save_captcha_sample(img, guess)
        _save_captcha_stats()
        log.info("    CAPTCHA auto-solved in %d attempt(s) -> %r", attempt, guess)
        return True
    log.warning("    CAPTCHA not solved within %d attempts.", CAPTCHA_MAX_TRIES)
    return False


def _debug_dump(page, tag: str) -> None:
    """Save a screenshot + visible page text so we can locate the OTP element.

    Only runs when BOT_DEBUG is set. Files land in output/debug/ and are named
    after the account so the screenshot and text line up.
    """
    import pathlib
    import re

    safe = re.sub(r"[^A-Za-z0-9._-]", "_", tag)[:60] or "row"
    out = pathlib.Path("output") / "debug"
    out.mkdir(parents=True, exist_ok=True)
    try:
        page.screenshot(path=str(out / f"{safe}.png"), full_page=True)
    except Exception:  # noqa: BLE001 - best-effort
        pass
    try:
        text = page.locator("body").inner_text(timeout=5_000)
        (out / f"{safe}.txt").write_text(text, encoding="utf-8")
    except Exception:  # noqa: BLE001 - best-effort
        pass
    try:
        (out / f"{safe}.html").write_text(page.content(), encoding="utf-8")
    except Exception:  # noqa: BLE001 - best-effort
        pass


def _extract_otp(text: str) -> str:
    """Pull the delivery OTP out of a cell like 'OutForDelivery (OTP: 7149)'.

    Returns the digit code, or '' when there is no OTP (e.g. 'Delivered' or a
    still-'In process' booking that has not gone out for delivery yet).
    """
    import re

    if not text:
        return ""
    # Preferred: an explicit 'OTP: 7149' token.
    m = re.search(r"OTP[:\s)]*?(\d{3,8})", text, re.IGNORECASE)
    if m:
        return m.group(1)
    low = text.lower()
    # Plain 'Delivered' (not 'OutForDelivery') has no OTP.
    if "deliver" in low and "outfor" not in low and "out for" not in low:
        return ""
    # Fallback: a standalone digit group near a delivery status.
    m2 = re.search(r"\b(\d{3,8})\b", text)
    return m2.group(1) if m2 else ""


def _capture_result(page) -> tuple[str, str]:
    """Read the OTP/Status and Order Ref No of the LATEST booking.

    On the booking-history grid the newest booking is the first data row. Its
    Status cell holds the delivery OTP while the booking is pending, or
    "Delivered" once the cylinder has been delivered. Returns
    ``(status_or_otp, order_ref_no)``.
    """
    try:
        table = page.locator(SEL_BOOKING_TABLE).first
        row = table.locator(BOOKING_ROW).first
        try:
            # Wait for the newest booking ROW to be ATTACHED (not "visible"): the
            # gridviewScroll plugin can wrap/hide the <table>, so requiring
            # "visible" times out even when the OTP is on screen. Shorter timeout
            # here since _open_booking_history already waited for the row.
            row.wait_for(state="attached", timeout=20_000)
        except PlaywrightTimeoutError:
            return ("NO_BOOKING_TABLE", "")

        cells = row.locator("td")
        n = cells.count()

        def _cell(i: int) -> str:
            if i >= n:
                return ""
            # text_content reads the text even if the element is plugin-hidden
            # (inner_text returns "" for display:none).
            text = (cells.nth(i).text_content(timeout=5_000) or "").strip()
            return " ".join(text.split())  # collapse whitespace/newlines

        status = _cell(STATUS_COL_IDX)
        delivery = _cell(DELIVERY_COL_IDX)
        order_ref = _cell(ORDERREF_COL_IDX)
        # The OTP, when present, sits in the delivery cell as
        # "OutForDelivery (OTP: 7149)". Prefer that; fall back to the Status
        # cell. If there's genuinely no OTP yet, return the human-readable
        # status so the column explains why (e.g. "Delivered" / "In process").
        otp = _extract_otp(delivery) or _extract_otp(status)
        if otp:
            return (otp, order_ref)
        return (delivery or status or "EMPTY_STATUS", order_ref)
    except Exception as e:  # noqa: BLE001 - best-effort
        return (f"CAPTURE_ERROR: {e}", "")


def _txt(value) -> str:
    """Coerce a cell value to a stripped string ('' for None)."""
    return "" if value is None else str(value).strip()


# Result values that represent a technical failure worth retrying on re-run
# (as opposed to a real status like an OTP, 'Delivered', or 'In process').
_RETRY_MARKERS = (
    "NO_BOOKING_TABLE", "NO_BOOKINGS", "EMPTY_STATUS", "CAPTURE_ERROR", "ERROR",
    "CAPTCHA_FAILED",
)


def _needs_retry(value) -> bool:
    """True if a previously recorded result should be attempted again."""
    s = _txt(value).upper()
    return (not s) or any(s.startswith(m) for m in _RETRY_MARKERS)


def _classify_inputs(input_dir: str):
    """Locate the master credentials file and the targets (consumer-ID list).

    The master is the workbook with many columns (>=5, i.e. it holds the
    credentials); the targets file is a single-column list of consumer IDs. If
    several of either exist, the most recently modified one wins.
    """
    xlsx = [
        p
        for p in pathlib.Path(input_dir).glob("*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not xlsx:
        raise FileNotFoundError(f"No .xlsx files found in {input_dir!r}.")

    masters, lists = [], []
    for p in xlsx:
        wb = load_workbook(p, read_only=True)
        max_col = wb.active.max_column or 1
        wb.close()
        (masters if max_col >= 3 else lists).append(p)

    if not masters:
        raise FileNotFoundError(
            "No master credentials file (a workbook with >=3 columns) found in input/."
        )
    master = max(masters, key=lambda p: p.stat().st_mtime)
    if lists:
        targets = max(lists, key=lambda p: p.stat().st_mtime)
    else:
        # Single-file mode: process all consumer IDs found in the master itself.
        targets = master
    return master, targets


def _load_targets(path) -> list[str]:
    """Return the ordered consumer IDs (as strings) from the targets workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    col = column_index_from_string(TARGETS_CONNO_COL)
    ids: list[str] = []
    for row in ws.iter_rows(min_row=START_ROW, min_col=col, max_col=col, values_only=True):
        v = row[0]
        if v not in (None, ""):
            ids.append(_txt(v))
    wb.close()
    return ids


def _colmap_for(path):
    """If a workbook's header row has email + password columns, return a colmap
    {id, email, pwd, out} of 1-based column indices; else None."""
    try:
        wb = load_workbook(path, read_only=True)
        header = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True), ())
        wb.close()
    except Exception:  # noqa: BLE001
        return None
    hdr = [_txt(h).lower() for h in header]
    email_i = next((i for i, h in enumerate(hdr, 1) if "email" in h), None)
    pwd_i = next((i for i, h in enumerate(hdr, 1) if h in ("pwd", "password", "pass")), None)
    if not (email_i and pwd_i):
        return None
    id_i = next((i for i, h in enumerate(hdr, 1)
                 if "consumer" in h or h in ("conno", "id", "consumerno")), 1)
    out_i = next((i for i, h in enumerate(hdr, 1) if "otp" in h or "out" in h), None)
    return {"id": id_i, "email": email_i, "pwd": pwd_i, "out": out_i}


def _detect_self_contained(input_dir: str):
    """Return (path, colmap) for the self-contained workbook to process.

    If BOT_FILE is set, that exact file is used; otherwise the newest input
    workbook whose header carries email + password columns. Returns None if no
    such file exists (then the master/targets mode is used).
    """
    if FILE_OVERRIDE:
        p = pathlib.Path(FILE_OVERRIDE)
        if not p.is_absolute():
            p = pathlib.Path(input_dir) / FILE_OVERRIDE
        cm = _colmap_for(p) if p.exists() else None
        if cm:
            return (p, cm)
    best = None
    for p in pathlib.Path(input_dir).glob("*.xlsx"):
        if not p.is_file() or p.name.startswith("~$"):
            continue
        cm = _colmap_for(p)
        if not cm:
            continue
        if best is None or p.stat().st_mtime > best[0].stat().st_mtime:
            best = (p, cm)
    return best


def _load_self_contained(path, colmap, start_row: int, end_row: int):
    """Load a self-contained workbook (writable) and read its account rows.

    Returns ``(workbook, worksheet, ordered_ids, info, out_col_idx)``. Rows
    outside the 1-based inclusive [start_row, end_row] window are skipped
    (0 = unbounded on that end). ``info[id] = {row, email, password}``.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb.active
    id_c, em_c, pw_c = colmap["id"], colmap["email"], colmap["pwd"]
    out_c = colmap["out"] or ((ws.max_column or 0) + 1)
    lo = start_row if start_row else 2
    hi = end_row if end_row else ws.max_row
    ordered: list[str] = []
    info: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, id_c).value
        if cid in (None, "") or r < lo or r > hi:
            continue
        key = _txt(cid)
        if key in info:
            continue
        info[key] = {
            "row": r,
            "email": _txt(ws.cell(r, em_c).value),
            "password": _txt(ws.cell(r, pw_c).value),
        }
        ordered.append(key)
    return wb, ws, ordered, info, out_c


def _load_master(path):
    """Open the master workbook (writable) and index it by consumer ID.

    Returns ``(workbook, worksheet, info)`` where ``info[consumer_id]`` is a
    dict ``{"row": int, "email": str, "password": str}``. The first occurrence
    of a consumer ID wins.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb[MASTER_SHEET] if MASTER_SHEET in wb.sheetnames else wb.active
    a = column_index_from_string(MASTER_CONNO_COL)
    c = column_index_from_string(MASTER_EMAIL_COL)
    e = column_index_from_string(MASTER_PWD_COL)
    info: dict[str, dict] = {}
    for r in range(1, ws.max_row + 1):
        cid = ws.cell(row=r, column=a).value
        if cid in (None, ""):
            continue
        key = _txt(cid)
        if key in info:
            continue
        info[key] = {
            "row": r,
            "email": _txt(ws.cell(row=r, column=c).value),
            "password": _txt(ws.cell(row=r, column=e).value),
        }
    return wb, ws, info


def _first_existing(paths: list[str]) -> str | None:
    for p in paths:
        if p and pathlib.Path(p).exists():
            return p
    return None


# Detected locations for the non-channel Chromium browsers.
_LOCALAPPDATA = os.environ.get("LOCALAPPDATA", "")
_PF = os.environ.get("ProgramFiles", r"C:\Program Files")
_PF86 = os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")
_BRAVE_PATHS = [
    rf"{_PF}\BraveSoftware\Brave-Browser\Application\brave.exe",
    rf"{_PF86}\BraveSoftware\Brave-Browser\Application\brave.exe",
    rf"{_LOCALAPPDATA}\BraveSoftware\Brave-Browser\Application\brave.exe",
]
_OPERA_PATHS = [
    rf"{_LOCALAPPDATA}\Programs\Opera GX\opera.exe",
    rf"{_LOCALAPPDATA}\Programs\Opera\opera.exe",
]


def _launch_browser(pw):
    """Launch the browser selected by BOT_BROWSER (default Chrome).

    Chrome/Edge use Playwright channels; Brave/Opera launch by executable path.
    Each account already runs in its own fresh context, so we don't need an
    explicit incognito flag for isolation.
    """
    log = logging.getLogger("bot")
    if BROWSER in ("edge", "msedge"):
        log.info("Browser: Microsoft Edge")
        return pw.chromium.launch(headless=HEADLESS, channel="msedge")
    if BROWSER == "brave":
        path = _first_existing(_BRAVE_PATHS)
        if path:
            log.info("Browser: Brave (%s)", path)
            return pw.chromium.launch(headless=HEADLESS, executable_path=path)
        log.warning("Brave not found; falling back to Chrome.")
    if BROWSER == "opera":
        path = _first_existing(_OPERA_PATHS)
        if path:
            log.info("Browser: Opera (%s)", path)
            return pw.chromium.launch(headless=HEADLESS, executable_path=path)
        log.warning("Opera not found; falling back to Chrome.")
    log.info("Browser: Google Chrome")
    return pw.chromium.launch(headless=HEADLESS, channel="chrome")


def _dismiss_ok(page, timeout: int = 4_000) -> None:
    """Best-effort: dismiss the post-login 'OK' confirmation modal if present."""
    try:
        ok = page.get_by_role("button", name="OK").first
        ok.wait_for(state="visible", timeout=timeout)
        ok.click()
    except Exception:  # noqa: BLE001 - best-effort
        pass


def _open_booking_history(page) -> bool:
    """Navigate to the booking grid, re-clicking the menu link if it's slow.

    We deliberately DO NOT reload the page. Reloading this ASP.NET portal
    resubmits the console form and re-pops the post-login 'OK' modal, which
    strands the run on the customer console. Instead, on each attempt we dismiss
    any stray OK modal and re-click the left-nav "View Cylinder Booking history"
    link -- a clean forward navigation. Returns True if the grid becomes visible.
    """
    for attempt in range(1, 3):
        _dismiss_ok(page, timeout=2_000)
        try:
            link = page.get_by_role(
                "link", name="View Cylinder Booking history"
            ).first
            link.wait_for(state="visible", timeout=15_000)
            link.click()
        except Exception:  # noqa: BLE001 - best-effort
            pass
        try:
            # Wait for a booking data ROW to be ATTACHED (present in the DOM),
            # not "visible": the gridviewScroll plugin wraps/hides the original
            # <table>, so the OTP can be on screen while the element reports
            # not-visible. "attached" detects the data either way.
            page.locator(f"{SEL_BOOKING_TABLE} {BOOKING_ROW}").first.wait_for(
                state="attached", timeout=45_000
            )
            return True
        except PlaywrightTimeoutError:
            logging.getLogger("bot").info(
                "  booking grid not ready (attempt %d/2) -> re-clicking menu...",
                attempt,
            )
    return False


def _process_row(browser, email: str, password: str) -> tuple[str, str]:
    """Run the recorded flow once for a single row; return the result string."""
    context = browser.new_context()
    context.set_default_timeout(ELEMENT_TIMEOUT_MS)
    context.set_default_navigation_timeout(NAV_TIMEOUT_MS)
    # Auto-accept any native JS dialogs so they never block the run.
    page = context.new_page()
    page.on("dialog", lambda d: d.accept())
    try:
        page.goto(LOGIN_URL)

        # 1. Email.
        email_field = page.locator(SEL_EMAIL).first
        email_field.wait_for(state="visible")
        email_field.fill(email)

        # 2. Solve the CAPTCHA -- automatically via OCR, or wait for a human.
        if AUTO_CAPTCHA:
            try:
                solved = _auto_solve_captcha(page)
            except Exception as e:  # noqa: BLE001 - OCR missing/broken -> human
                logging.getLogger("bot").warning(
                    "    auto-CAPTCHA unavailable (%s); waiting for a human.", e)
                _wait_for_captcha_solved(page, email)
                solved = True
            if not solved:
                return ("CAPTCHA_FAILED", "")
        else:
            _wait_for_captcha_solved(page, email)

        # 3. Password (appears after the CAPTCHA is accepted).
        pwd_field = page.locator(SEL_PASSWORD).first
        pwd_field.wait_for(state="visible")
        pwd_field.fill(password)

        # 4. Submit.
        page.locator(SEL_LOGIN_BTN).first.click()

        # 5. Best-effort: dismiss the post-login "OK" confirmation modal.
        _dismiss_ok(page, timeout=5_000)

        # Let the server finish setting up the session before we request the
        # heavy booking-history page. This human-like pause is the main defence
        # against the booking grid failing to load right after a fast login.
        if SETTLE_MS > 0:
            time.sleep(SETTLE_MS / 1000.0)

        # 6. Open the booking-history grid, re-clicking the menu link if the
        #    flaky page is slow (no reload -> never bounces back to the console).
        _open_booking_history(page)

        if DEBUG_DUMP:
            _debug_dump(page, email)

        return _capture_result(page)
    finally:
        try:
            context.close()
        except Exception:  # noqa: BLE001 - best-effort cleanup
            pass


def _safe_save(wb, path) -> bool:
    """Save a workbook, but never crash the run if the file is locked/open.

    If the target is open in Excel (a Windows lock), we log a warning and keep
    going -- progress is not lost, the file just isn't updated until it's closed.
    """
    try:
        wb.save(str(path))
        return True
    except (PermissionError, OSError) as e:
        logging.getLogger("bot").warning(
            "    could not save %s (is it open in Excel?): %s", path, e)
        return False


class _Pacer:
    """Adaptive inter-login delay that rides the portal's per-IP throttle.

    Strategy (a throttle-aware AIMD):
      * a booking-page **failure** (NO_BOOKING_TABLE / ERROR -- the throttle
        signature) multiplies the gap by 1.6 (back off fast);
      * every few **clean successes** shave 1s off the gap (probe faster);
      * a sustained failure **streak** triggers a one-off cooldown sleep so the
        server can recover, then resumes at a calmer rate.

    All bounded by BOT_MIN_BETWEEN_MS .. BOT_MAX_BETWEEN_MS. With BOT_ADAPTIVE=0
    it just sleeps a fixed BOT_BETWEEN_MS.
    """

    def __init__(self):
        self.between = float(BETWEEN_MS)
        self.fail_streak = 0
        self.ok_streak = 0
        self._log = logging.getLogger("bot")

    def wait(self, first: bool) -> None:
        """Sleep the current gap before a login (nothing before the very first)."""
        if first:
            return
        gap = self.between if ADAPTIVE else float(BETWEEN_MS)
        if gap > 0:
            time.sleep(gap / 1000.0)

    def record(self, ok: bool) -> None:
        """Update the gap from the latest booking-page outcome."""
        if not ADAPTIVE:
            return
        if ok:
            self.fail_streak = 0
            self.ok_streak += 1
            if self.ok_streak >= 3 and self.between > MIN_BETWEEN_MS:
                self.between = max(MIN_BETWEEN_MS, self.between - 1000)
                self.ok_streak = 0
                self._log.info("    [pacing] healthy -> gap %.1fs", self.between / 1000.0)
        else:
            self.ok_streak = 0
            self.fail_streak += 1
            new = min(MAX_BETWEEN_MS, self.between * 1.6)
            if new != self.between:
                self.between = new
                self._log.info("    [pacing] throttle? backing off -> gap %.1fs",
                               self.between / 1000.0)

    def maybe_cooldown(self) -> None:
        """If failures are piling up, pause to let the portal's throttle reset."""
        if ADAPTIVE and self.fail_streak >= COOLDOWN_FAILS and COOLDOWN_MS > 0:
            self._log.warning(
                "    [pacing] %d failures in a row -> cooling down %.0fs to let "
                "the portal recover...", self.fail_streak, COOLDOWN_MS / 1000.0)
            time.sleep(COOLDOWN_MS / 1000.0)
            self.fail_streak = 0
            self.between = min(MAX_BETWEEN_MS, max(self.between, float(BETWEEN_MS)))


def _write_captured_otps(res_ws, out_dir: str) -> int:
    """Write the clean 3-column Captured_OTPs.xlsx (Consumer ID | otp | value)."""
    cap = Workbook()
    cw = cap.active
    cw.title = "OTPs"
    found = 0
    for r in range(2, res_ws.max_row + 1):
        cid = _txt(res_ws.cell(r, 1).value)
        val = _txt(res_ws.cell(r, 2).value)
        if cid and re.fullmatch(r"\d{3,8}", val):
            cw.append([cid, "otp", val])
            found += 1
    _safe_save(cap, pathlib.Path(out_dir) / "Captured_OTPs.xlsx")
    return found


def main() -> int:
    _setup_logging()
    log = logging.getLogger("bot")

    out_dir = pathlib.Path(OUTPUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    sc = _detect_self_contained(INPUT_DIR)
    if sc is not None:
        # --- Self-contained mode: one workbook holds IDs + email + password. ---
        sc_path, colmap = sc
        master_wb, master_ws, target_ids, info, otp_col_idx = _load_self_contained(
            sc_path, colmap, START_ROW, END_ROW)
        otp_col_letter = get_column_letter(otp_col_idx)
        window = (f"rows {START_ROW or 2}..{END_ROW}" if END_ROW
                  else (f"rows {START_ROW}+" if START_ROW else "all rows"))
        log.info("Self-contained file: %s -> %d account(s) (%s); OTP -> column %s.",
                 sc_path.name, len(target_ids), window, otp_col_letter)
        master_out = out_dir / f"{sc_path.stem}.result.xlsx"
        otp_out = out_dir / f"OTP_results_{sc_path.stem}.xlsx"
    else:
        # --- Master + targets mode. ---
        master_path, targets_path = _classify_inputs(INPUT_DIR)
        log.info("Master credentials file : %s", master_path.name)
        log.info("Targets (consumer IDs)  : %s", targets_path.name)
        target_ids = _load_targets(targets_path)
        log.info("Loaded %d target consumer ID(s).", len(target_ids))
        if MASTER_INDEX and pathlib.Path(MASTER_INDEX).exists():
            info = json.loads(pathlib.Path(MASTER_INDEX).read_text(encoding="utf-8"))
            master_wb = master_ws = None
            otp_col_idx = None
            otp_col_letter = "-"
            log.info("Loaded prebuilt master index (%d ids); shard writes only its OTP file.",
                     len(info))
        else:
            log.info("Loading master workbook (this can take ~20s on the big file)...")
            master_wb, master_ws, info = _load_master(master_path)
            log.info("Master indexed: %d unique consumer IDs.", len(info))
            otp_col_idx = (master_ws.max_column or 0) + 1
            otp_col_letter = get_column_letter(otp_col_idx)
            log.info("OTP -> master column %s (new last column).", otp_col_letter)
        master_out = out_dir / f"{master_path.stem}.result{OUT_SUFFIX}.xlsx"
        otp_out = out_dir / f"OTP_results{OUT_SUFFIX}.xlsx"

    # Sharding + MAX_ROWS apply to the final target list in either mode.
    if SHARD_COUNT > 1:
        n = len(target_ids)
        per = -(-n // SHARD_COUNT)  # ceil division
        start = SHARD_INDEX * per
        target_ids = target_ids[start:start + per]
        log.info("Shard %d/%d -> %d ids.", SHARD_INDEX + 1, SHARD_COUNT, len(target_ids))
    if MAX_ROWS is not None:
        target_ids = target_ids[:MAX_ROWS]
        log.info("BOT_MAX_ROWS set -> processing only the first %d.", len(target_ids))

    # Separate OTP results workbook (Consumer ID | OTP). Resume from a prior run
    # if one exists, so we only re-solve CAPTCHAs for what's left / what failed.
    if otp_out.exists():
        res_wb = load_workbook(str(otp_out))
        res_ws = res_wb.active
        log.info("Resuming from existing %s.", otp_out.name)
    else:
        res_wb = Workbook()
        res_ws = res_wb.active
        res_ws.title = "OTP"
        res_ws.append(["Consumer ID", "OTP"])
    # Index existing result rows by consumer ID for in-place upsert.
    res_rows: dict[str, int] = {}
    for r in range(2, res_ws.max_row + 1):
        key = _txt(res_ws.cell(row=r, column=1).value)
        if key:
            res_rows[key] = r

    def _record(cid: str, otp: str) -> None:
        if cid in res_rows:
            res_ws.cell(row=res_rows[cid], column=2, value=otp)
        else:
            res_ws.append([cid, otp])
            res_rows[cid] = res_ws.max_row

    processed = 0   # accounts actually processed across all passes
    skipped = 0
    total = len(target_ids)
    pacer = _Pacer()

    def _remaining() -> int:
        """How many target accounts still need work (failed or not yet done)."""
        c = 0
        for cid in target_ids:
            rec = info.get(cid)
            if rec is None or not rec.get("email"):
                continue  # NOT_IN_MASTER -- not retryable
            prev = res_ws.cell(row=res_rows[cid], column=2).value if cid in res_rows else None
            if prev is None or _needs_retry(prev):
                c += 1
        return c

    max_passes = MAX_PASSES if AUTO_RETRY else 1
    interrupted = False
    dry_passes = 0
    with sync_playwright() as pw:
        browser = _launch_browser(pw)
        try:
            for pass_no in range(1, max_passes + 1):
                pass_done = 0   # processed this pass
                pass_ok = 0     # newly captured this pass
                log.info("==== Pass %d/%d -- %d account(s) still to do ====",
                         pass_no, max_passes, _remaining())
                for i, cid in enumerate(target_ids, 1):
                    rec = info.get(cid)
                    if rec is None or not rec.get("email"):
                        if pass_no == 1:
                            log.warning("[%d/%d] Consumer %s: not in master / no email -> skipped.",
                                        i, total, cid)
                            _record(cid, "NOT_IN_MASTER")
                            _safe_save(res_wb, otp_out)
                        continue

                    prev = res_ws.cell(row=res_rows[cid], column=2).value if cid in res_rows else None
                    if prev is not None and not _needs_retry(prev):
                        if master_ws is not None:
                            master_ws.cell(row=rec["row"], column=otp_col_idx, value=prev)
                        if pass_no == 1:
                            skipped += 1
                        continue

                    # Adaptive delay before each login (none before the very first).
                    pacer.wait(first=(processed == 0))

                    log.info("[p%d %d/%d] Consumer %s -> signing in as %s",
                             pass_no, i, total, cid, rec["email"])
                    try:
                        otp, order_ref = _process_row(browser, rec["email"], rec["password"])
                    except KeyboardInterrupt:
                        log.warning("Interrupted by user. Saving progress and exiting.")
                        interrupted = True
                        break
                    except Exception as ex:  # noqa: BLE001 - keep going on the next account
                        log.exception("Consumer %s failed: %s", cid, ex)
                        otp, order_ref = (f"ERROR: {ex}", "")

                    if master_ws is not None:
                        master_ws.cell(row=rec["row"], column=otp_col_idx, value=otp)
                    _record(cid, otp)
                    processed += 1
                    pass_done += 1

                    ok = not _needs_retry(otp)
                    if ok:
                        pass_ok += 1
                    pacer.record(ok)
                    log.info("[p%d %d/%d] Consumer %s -> OTP=%r  (gap %.1fs)",
                             pass_no, i, total, cid, otp, pacer.between / 1000.0)

                    _safe_save(res_wb, otp_out)
                    if master_ws is not None and processed % 10 == 0:
                        _safe_save(master_wb, master_out)
                    if not ok:
                        pacer.maybe_cooldown()

                rem = _remaining()
                log.info("==== Pass %d done: %d processed, %d captured, %d still failing ====",
                         pass_no, pass_done, pass_ok, rem)
                if interrupted or rem == 0 or not AUTO_RETRY:
                    break
                if pass_ok == 0:
                    # No progress -- likely the portal is throttled/down. Cool
                    # down to let it recover and try again; give up only after
                    # DRY_STOP dry passes in a row.
                    dry_passes += 1
                    if dry_passes >= DRY_STOP:
                        log.warning("No progress for %d passes -> stopping (%d still failing). "
                                    "Re-run later to retry them.", dry_passes, rem)
                        break
                    log.warning("No progress (dry pass %d/%d) -> cooling down %.0fs to let the "
                                "portal recover, then retrying...", dry_passes, DRY_STOP,
                                COOLDOWN_MS / 1000.0)
                    if COOLDOWN_MS > 0:
                        time.sleep(COOLDOWN_MS / 1000.0)
                    pacer.between = float(BETWEEN_MS)  # reset the pace after a cooldown
                    continue
                dry_passes = 0
                breather = min(COOLDOWN_MS, 60_000) / 1000.0
                if breather > 0:
                    log.info("Retry pass next; pausing %.0fs first.", breather)
                    time.sleep(breather)
        finally:
            try:
                browser.close()
            except Exception:  # noqa: BLE001
                pass

    _safe_save(res_wb, otp_out)
    if master_ws is not None:
        _safe_save(master_wb, master_out)
        # Auto-generate the clean 3-column OTP file (normal, non-shard runs).
        ncap = _write_captured_otps(res_ws, OUTPUT_DIR)
        log.info("Captured OTPs (clean 3-col) -> %d in output/Captured_OTPs.xlsx", ncap)
    log.info("Done. %d processed, %d already-had, %d total target(s).",
             processed, skipped, total)
    log.info("OTP list  -> %s", otp_out)
    if master_ws is not None:
        log.info("Master+OTP (column %s) -> %s", otp_col_letter, master_out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
