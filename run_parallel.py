"""Run the OTP bot across several browsers in parallel, each on a unique slice.

Usage (from the project root, with the venv python):

    .venv/Scripts/python.exe run_parallel.py chrome,brave,edge

Each browser gets a distinct, CONTIGUOUS slice of the target consumer IDs, so no
ID is ever processed by two browsers and no login is reused. Each browser writes
its own ``output/OTP_results_<browser>.xlsx``; when all finish they are merged
into ``output/OTP_results.xlsx`` and ``output/Captured_OTPs.xlsx``.

WARNING -- read before relying on this for speed:
All browsers here share this machine's ONE IP. The My HPGas portal throttles
rapid logins per IP (we measured ~5-8s between logins as the safe rate), so
several browsers hammering it together can RAISE the failure rate rather than
throughput. The reliable way to parallelize is across DEVICES (different IPs):
give each device a shard with BOT_SHARD_INDEX / BOT_SHARD_COUNT. Use this
launcher to TEST whether one-IP parallelism actually helps your situation.

Env knobs: BOT_BROWSERS (default "chrome,brave,edge"), BOT_STAGGER_MS (delay
between launching each browser, default 2000), plus the usual BOT_SETTLE_MS /
BOT_BETWEEN_MS / BOT_AUTO_CAPTCHA / BOT_HEADLESS which pass through to each shard.
"""
from __future__ import annotations

import json
import os
import pathlib
import re
import shutil
import subprocess
import sys
import time

from openpyxl import Workbook, load_workbook

import run_bot as rb

_OTP_RE = re.compile(r"\d{3,8}")


def _is_otp(v: str) -> bool:
    return bool(_OTP_RE.fullmatch(v or ""))


def main() -> int:
    browsers = (sys.argv[1] if len(sys.argv) > 1
                else os.environ.get("BOT_BROWSERS", "chrome,brave,edge"))
    browsers = [b.strip().lower() for b in browsers.split(",") if b.strip()]
    n = len(browsers)
    if n == 0:
        print("No browsers specified."); return 1
    stagger_s = float(os.environ.get("BOT_STAGGER_MS", "2000")) / 1000.0

    out = pathlib.Path(rb.OUTPUT_DIR)
    out.mkdir(parents=True, exist_ok=True)
    main_otp = out / "OTP_results.xlsx"

    # 1. Build the credentials index ONCE so shards don't each load the master.
    master_path, targets_path = rb._classify_inputs(rb.INPUT_DIR)
    print(f"[parallel] master={master_path.name}  targets={targets_path.name}")
    print("[parallel] building shared credential index (~20s)...")
    _wb, _ws, info = rb._load_master(master_path)
    idx_path = out / "_master_index.json"
    idx_path.write_text(json.dumps(info), encoding="utf-8")
    print(f"[parallel] index: {len(info)} ids -> {idx_path.name}")
    print(f"[parallel] launching {n} browser(s) {browsers}; stagger {stagger_s}s")

    # 2. Launch one subprocess per browser, each with its own shard + output.
    procs = []
    for i, br in enumerate(browsers):
        suffix = f"_{br}"
        shard_otp = out / f"OTP_results{suffix}.xlsx"
        # Seed with prior progress so the shard resumes (skips already-captured).
        if main_otp.exists():
            shutil.copyfile(main_otp, shard_otp)
        env = dict(os.environ)
        env.update({
            "BOT_BROWSER": br,
            "BOT_SHARD_INDEX": str(i),
            "BOT_SHARD_COUNT": str(n),
            "BOT_OUT_SUFFIX": suffix,
            "BOT_MASTER_INDEX": str(idx_path),
            "BOT_AUTO_CAPTCHA": env.get("BOT_AUTO_CAPTCHA", "1"),
            "BOT_SETTLE_MS": env.get("BOT_SETTLE_MS", "5000"),
            "BOT_BETWEEN_MS": env.get("BOT_BETWEEN_MS", "6000"),
        })
        logf = open(out / f"parallel_{br}.log", "w", encoding="utf-8")
        p = subprocess.Popen([sys.executable, "run_bot.py"], env=env,
                             stdout=logf, stderr=subprocess.STDOUT)
        procs.append((br, p, logf))
        print(f"[parallel] started {br} (shard {i + 1}/{n}) -> output/parallel_{br}.log")
        time.sleep(stagger_s)

    # 3. Wait for all browsers to finish.
    for br, p, logf in procs:
        p.wait()
        logf.close()
        print(f"[parallel] {br} finished (exit {p.returncode}).")

    # 4. Merge: union all per-shard files (+ any prior main), preferring a real OTP.
    merged: dict[str, str] = {}
    sources = [out / f"OTP_results_{br}.xlsx" for br in browsers] + [main_otp]
    for f in sources:
        if not f.exists():
            continue
        ws = load_workbook(str(f), data_only=True).active
        for r in range(2, ws.max_row + 1):
            cid = rb._txt(ws.cell(r, 1).value)
            val = rb._txt(ws.cell(r, 2).value)
            if not cid:
                continue
            if cid not in merged or (_is_otp(val) and not _is_otp(merged[cid])):
                merged[cid] = val

    res = Workbook(); rw = res.active; rw.title = "OTP"
    rw.append(["Consumer ID", "OTP"])
    for cid, val in merged.items():
        rw.append([cid, val])
    res.save(str(main_otp))

    cap = Workbook(); cw = cap.active; cw.title = "OTPs"
    found = 0
    for cid, val in merged.items():
        if _is_otp(val):
            cw.append([cid, "otp", val]); found += 1
    cap.save(str(out / "Captured_OTPs.xlsx"))

    print(f"[parallel] merged {len(merged)} results; {found} OTPs captured.")
    print(f"[parallel] -> {main_otp}")
    print(f"[parallel] -> {out / 'Captured_OTPs.xlsx'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
