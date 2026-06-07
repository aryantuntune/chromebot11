"""Browser-free unit tests for :mod:`src.excel_io`.

These tests build real ``.xlsx`` files in pytest's ``tmp_path`` fixture, so
they exercise the actual openpyxl read/write paths without needing Playwright.
"""

from __future__ import annotations

import os
import time

import pytest
from openpyxl import Workbook, load_workbook

from src.config import ExcelConfig
from src.excel_io import (
    RowData,
    find_input_file,
    load_workbook_and_rows,
    save_workbook,
    write_result,
)


def _make_excel_config(input_dir, output_dir) -> ExcelConfig:
    """Build an ExcelConfig matching the shared contract."""
    return ExcelConfig(
        input_dir=str(input_dir),
        output_dir=str(output_dir),
        sheet_name=None,      # use the active sheet
        header_row=1,
        first_data_row=2,
        email_col="A",
        password_col="B",
        result_col="C",
    )


def _build_workbook(path, data_rows, header=("email", "password", "result")):
    """Create a workbook at ``path`` with a header then ``data_rows``.

    ``data_rows`` is a list of (email, password, result) tuples; any element may
    be ``None`` to leave a cell blank.
    """
    wb = Workbook()
    ws = wb.active
    if header is not None:
        ws.append(list(header))
    for row in data_rows:
        ws.append(list(row))
    wb.save(str(path))
    return path


def test_find_input_file_picks_newest_and_ignores_temp(tmp_path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()

    old = input_dir / "old.xlsx"
    new = input_dir / "new.xlsx"
    _build_workbook(old, [("a@x.com", "pw", None)])
    _build_workbook(new, [("b@x.com", "pw", None)])

    # Force a clear modification-time ordering: old is older than new.
    old_time = time.time() - 1000
    os.utime(old, (old_time, old_time))

    # Excel temp/lock file -- must be ignored even if it's the newest.
    temp = input_dir / "~$new.xlsx"
    temp.write_bytes(b"not a real workbook")

    picked = find_input_file(str(input_dir))
    assert picked.name == "new.xlsx"


def test_find_input_file_raises_when_empty(tmp_path):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    with pytest.raises(FileNotFoundError):
        find_input_file(str(input_dir))


def test_load_rows_reads_credentials_and_skips_empty_email(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    path = input_dir / "creds.xlsx"

    _build_workbook(
        path,
        [
            ("alice@example.com", "secret1", None),
            (None, "orphan-password", None),   # empty email -> skipped
            ("", "blank-email", None),         # empty-string email -> skipped
            ("bob@example.com", "secret2", "old-result"),
        ],
    )

    cfg = _make_excel_config(input_dir, output_dir)
    wb, ws, rows = load_workbook_and_rows(path, cfg)

    assert [r.email for r in rows] == ["alice@example.com", "bob@example.com"]
    assert [r.password for r in rows] == ["secret1", "secret2"]
    # Worksheet rows: header=1, alice=2, (skip 3,4), bob=5.
    assert rows[0].row_number == 2
    assert rows[1].row_number == 5
    assert isinstance(rows[0], RowData)
    assert rows[1].result == "old-result"


def test_write_result_and_save_workbook_leaves_input_untouched(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    path = input_dir / "creds.xlsx"

    _build_workbook(path, [("alice@example.com", "secret1", None)])
    original_bytes = path.read_bytes()

    cfg = _make_excel_config(input_dir, output_dir)
    wb, ws, rows = load_workbook_and_rows(path, cfg)

    write_result(ws, cfg, rows[0].row_number, "HELLO MESSAGE")
    out_path = save_workbook(wb, path, str(output_dir))

    # Output file lives in the output directory with the stable derived name.
    assert out_path.parent == output_dir
    assert out_path.name == "creds.result.xlsx"
    assert out_path.is_file()

    # The result was written into column C (3rd column).
    out_wb = load_workbook(str(out_path))
    out_ws = out_wb.active
    assert out_ws.cell(row=rows[0].row_number, column=3).value == "HELLO MESSAGE"

    # The original input file is byte-for-byte unchanged.
    assert path.read_bytes() == original_bytes


def test_save_workbook_is_stable_across_incremental_saves(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    path = input_dir / "creds.xlsx"

    _build_workbook(
        path,
        [("alice@example.com", "s1", None), ("bob@example.com", "s2", None)],
    )

    cfg = _make_excel_config(input_dir, output_dir)
    wb, ws, rows = load_workbook_and_rows(path, cfg)

    write_result(ws, cfg, rows[0].row_number, "first")
    first_out = save_workbook(wb, path, str(output_dir))
    write_result(ws, cfg, rows[1].row_number, "second")
    second_out = save_workbook(wb, path, str(output_dir))

    # Same stable filename overwritten in place -> only one output file.
    assert first_out == second_out
    assert list(output_dir.glob("*.xlsx")) == [second_out]

    out_wb = load_workbook(str(second_out))
    out_ws = out_wb.active
    assert out_ws.cell(row=rows[0].row_number, column=3).value == "first"
    assert out_ws.cell(row=rows[1].row_number, column=3).value == "second"
