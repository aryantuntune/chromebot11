r"""Live dashboard for the HP Gas OTP bot.

Run in a separate terminal while supervisor.py (or run_bot.py) is active:
    .venv\Scripts\python.exe dashboard.py

Refreshes every 3 seconds. Press Ctrl+C to exit.
"""
from __future__ import annotations

import json
import pathlib
import re
import time
from datetime import datetime

from openpyxl import load_workbook
from rich import box
from rich.console import Console
from rich.layout import Layout
from rich.live import Live
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

import os

REFRESH_SEC  = 3
OUTPUT_DIR   = pathlib.Path("output")
OTP_RESULTS  = OUTPUT_DIR / "OTP_results.xlsx"
CAP_STATS    = OUTPUT_DIR / "captcha_stats.json"
SUP_STATUS   = OUTPUT_DIR / "supervisor_status.json"
# Same 1-based inclusive row window the supervisor/bot use (0 = unbounded).
START_ROW    = int(os.environ.get("BOT_START_ROW", "0"))
END_ROW      = int(os.environ.get("BOT_END_ROW",   "0"))

_RETRY_MARKERS = (
    "NO_BOOKING_TABLE", "NO_BOOKINGS", "EMPTY_STATUS",
    "CAPTURE_ERROR", "ERROR", "CAPTCHA_FAILED",
)


# ── Data loaders ─────────────────────────────────────────────────────────────

def _find_input_file():
    candidates = [
        p for p in pathlib.Path("input").glob("*.xlsx")
        if not p.name.startswith("~$")
    ]
    masters = []
    for p in candidates:
        try:
            wb = load_workbook(p, read_only=True)
            cols = wb.active.max_column or 1
            wb.close()
            if cols >= 3:
                masters.append(p)
        except Exception:
            pass
    if not masters:
        return None, []
    best = max(masters, key=lambda p: p.stat().st_mtime)
    try:
        wb = load_workbook(str(best), read_only=True, data_only=True)
        ws = wb.active
        lo = START_ROW if START_ROW else 2
        hi = END_ROW if END_ROW else ws.max_row
        ids = []
        for r in range(2, ws.max_row + 1):
            if r < lo or r > hi:
                continue
            v = ws.cell(r, 1).value
            if v not in (None, ""):
                ids.append(str(v).strip())
        wb.close()
        return best, ids
    except Exception:
        return best, []


def _load_results(target_set: set[str]):
    if not OTP_RESULTS.exists():
        return {}
    try:
        wb = load_workbook(str(OTP_RESULTS), read_only=True, data_only=True)
        ws = wb.active
        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            cid = str(row[0] or "").strip()
            val = str(row[1] or "").strip()
            if cid in target_set:
                data[cid] = val
        wb.close()
        return data
    except Exception:
        return {}


def _load_captcha_stats():
    try:
        return json.loads(CAP_STATS.read_text(encoding="utf-8")) if CAP_STATS.exists() else {}
    except Exception:
        return {}


def _load_supervisor_status():
    try:
        return json.loads(SUP_STATUS.read_text(encoding="utf-8")) if SUP_STATUS.exists() else {}
    except Exception:
        return {}


# ── Classify a result value ───────────────────────────────────────────────────

def _classify(val: str):
    """Return (category, display_text, style)."""
    if not val:
        return "missing", "— pending —", "dim"
    if re.fullmatch(r"\d{3,8}", val):
        return "captured", f"✔  {val}", "bold green"
    up = val.upper()
    if up.startswith("NO_BOOKING"):
        return "retry", "⚠  NO_BOOKING_TABLE", "yellow"
    if up.startswith("CAPTCHA"):
        return "retry", "⚠  CAPTCHA_FAILED", "yellow"
    if up.startswith("EMPTY"):
        return "retry", "⚠  EMPTY_STATUS", "yellow"
    if up.startswith("ERROR") or up.startswith("CAPTURE_ERROR"):
        return "retry", "✖  ERROR", "red"
    if up.startswith("NOT_IN"):
        return "other", "–  NOT_IN_MASTER", "dim"
    if "deliver" in up and "outfor" not in up and "out for" not in up:
        return "other", f"📦  {val[:20]}", "cyan"
    if "process" in up:
        return "other", f"⏳  {val[:20]}", "blue"
    return "other", val[:25], "dim"


# ── Build the display ─────────────────────────────────────────────────────────

def build_display(input_file, target_ids, results, cap_stats, sup_status) -> Layout:
    total = len(target_ids)
    now   = datetime.now().strftime("%H:%M:%S")

    # Tally
    counts = {"captured": 0, "retry": 0, "other": 0, "missing": 0}
    for cid in target_ids:
        cat, _, _ = _classify(results.get(cid, ""))
        counts[cat] += 1

    processed  = total - counts["missing"]
    pct        = processed / total * 100 if total else 0
    cap_pct    = counts["captured"] / total * 100 if total else 0

    layout = Layout()
    layout.split_column(
        Layout(name="header",  size=3),
        Layout(name="body"),
        Layout(name="footer",  size=3),
    )
    layout["body"].split_row(
        Layout(name="left",  ratio=2),
        Layout(name="right", ratio=3),
    )

    # ── Header ───────────────────────────────────────────────────────────────
    fname = input_file.name if input_file else "unknown"
    layout["header"].update(Panel(
        Text(f"  HP Gas OTP Bot  •  {fname}  •  {now}  •  Ctrl+C to exit",
             style="bold white on blue"),
        box=box.HEAVY,
    ))

    # ── Left panel: stats ─────────────────────────────────────────────────────
    bar_len    = 48
    filled     = int(cap_pct / 100 * bar_len)
    prog_bar   = "[green]" + "█" * filled + "[/green][dim]" + "░" * (bar_len - filled) + "[/dim]"

    summary = (
        f"[bold]File[/bold]      : [cyan]{fname}[/cyan]\n"
        f"[bold]Total[/bold]     : {total}\n\n"
        f"[bold green]Captured  : {counts['captured']:>4}[/bold green]  "
        f"([green]{cap_pct:.1f}%[/green])\n"
        f"[bold yellow]Retryable : {counts['retry']:>4}[/bold yellow]\n"
        f"[cyan]Other     : {counts['other']:>4}[/cyan]  (delivered / in-process)\n"
        f"[dim]Pending   : {counts['missing']:>4}[/dim]  (not yet attempted)\n\n"
        f"Processed : {processed}/{total}  ({pct:.1f}%)\n\n"
        f"{prog_bar}\n"
        f"[dim]OTP capture progress: {cap_pct:.1f}%[/dim]"
    )

    # Captcha block
    cap_block = ""
    if cap_stats:
        cap_block = (
            "\n\n[bold]CAPTCHA Solver[/bold]\n"
            f"  Attempts      : {cap_stats.get('attempts', '?')}\n"
            f"  Solved        : {cap_stats.get('successes', '?')}\n"
            f"  First-try     : {cap_stats.get('first_try_rate_pct', '?')}%\n"
            f"  Overall rate  : {cap_stats.get('overall_solve_rate_pct', '?')}%"
        )

    # Supervisor block
    sup_block = ""
    if sup_status:
        state = sup_status.get("state", "?")
        run   = sup_status.get("run", "?")
        stall = sup_status.get("stall_streak", 0)
        state_style = {
            "running": "bold green",
            "idle": "dim",
            "stuck": "bold red",
        }.get(state, "white")
        sup_block = (
            f"\n\n[bold]Supervisor[/bold]\n"
            f"  Run #         : {run}\n"
            f"  State         : [{state_style}]{state}[/{state_style}]\n"
            f"  Stall streak  : {stall}"
        )

    layout["left"].update(Panel(
        summary + cap_block + sup_block,
        title="[bold]Stats[/bold]",
        border_style="cyan",
    ))

    # ── Right panel: results table ────────────────────────────────────────────
    tbl = Table(
        "Consumer ID", "OTP / Status",
        box=box.SIMPLE_HEAD,
        show_header=True,
        header_style="bold white",
        expand=True,
        show_lines=False,
    )

    # Sort: captures first (most valuable), then retries, then other, then missing
    order = {"captured": 0, "retry": 1, "other": 2, "missing": 3}
    sorted_ids = sorted(
        target_ids,
        key=lambda cid: (order[_classify(results.get(cid, ""))[0]], cid)
    )

    shown = 0
    for cid in sorted_ids:
        if shown >= 60:
            break
        val = results.get(cid, "")
        _, disp, style = _classify(val)
        tbl.add_row(cid, Text(disp, style=style))
        shown += 1

    remaining_count = total - shown
    layout["right"].update(Panel(
        tbl,
        title=f"[bold]Account Results[/bold] [dim](showing {shown}/{total})[/dim]",
        border_style="green",
    ))

    # ── Footer ────────────────────────────────────────────────────────────────
    layout["footer"].update(Panel(
        f"[dim]Refresh: every {REFRESH_SEC}s  •  "
        f"Results: {OTP_RESULTS}  •  "
        f"Run supervisor.py to auto-retry all failures[/dim]",
        box=box.SIMPLE,
    ))

    return layout


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    console = Console()
    console.clear()
    input_file, target_ids = _find_input_file()

    with Live(console=console, refresh_per_second=1, screen=True) as live:
        try:
            while True:
                target_set  = set(target_ids)
                results     = _load_results(target_set)
                cap_stats   = _load_captcha_stats()
                sup_status  = _load_supervisor_status()
                live.update(build_display(input_file, target_ids, results, cap_stats, sup_status))
                time.sleep(REFRESH_SEC)
        except KeyboardInterrupt:
            pass

    console.print("\n[green]Dashboard closed.[/green]")


if __name__ == "__main__":
    main()
