"""Restore the 5 OTPs that were overwritten by duplicate bot processes."""
from openpyxl import load_workbook
import pathlib, re

# Known-correct OTPs from monitor log before processes clashed
KNOWN_GOOD = {
    "651155": "3696",
    "651188": "6249",
    "651190": "3396",
    "651201": "2484",
    "651227": "8179",
    "651249": "5412",
}

p = pathlib.Path("output/OTP_results.xlsx")
wb = load_workbook(str(p), data_only=False)
ws = wb.active

patched = 0
for row in ws.iter_rows(min_row=2):
    cid = str(row[0].value or "").strip()
    val = str(row[1].value or "").strip()
    if cid in KNOWN_GOOD:
        correct = KNOWN_GOOD[cid]
        if val != correct:
            print(f"  Patching {cid}: {val!r} -> {correct!r}")
            row[1].value = correct
            patched += 1
        else:
            print(f"  {cid}: already correct ({val})")

wb.save(str(p))
print(f"\nPatched {patched} row(s). Saved.")
