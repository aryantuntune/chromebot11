"""Overnight orchestrator — work through every self-contained input file, one
after another, fully hands-free, with LONG cooldowns when the portal throttles.

It runs `run_bot.py` per file (forcing that file + its row range), lets run_bot's
own pass/retry engine drain it, then moves to the next file. Files you DROP IN
while it's running are picked up on the next round. It keeps cycling until every
file's accounts are captured — or you stop it (Ctrl+C / close window). Re-running
resumes (already-captured accounts are skipped).

Run:  .venv/Scripts/python.exe run_overnight.py

Why overnight: this portal throttles hard during the day; it serves the booking
page far more reliably late night / early morning. Letting this loop run while
you sleep means it captures the easy windows and rests (cooldowns) during slow
ones, instead of you babysitting it.
"""
from __future__ import annotations

import os
import pathlib
import re
import subprocess
import sys
import time

from openpyxl import load_workbook

import run_bot as rb

# Per-file row ranges (1-based, inclusive). Files NOT listed are done in full.
# Add an entry here (or just drop the file in input/) for each new file.
RANGES = {
    "asu-330-0706 (1).xlsx": (247, 312),
}

# Settings handed to every run_bot invocation (override via your own env first).
OVERNIGHT_ENV = {
    "BOT_BROWSER": os.environ.get("BOT_BROWSER", "edge"),
    "BOT_AUTO_CAPTCHA": "1",
    "BOT_ADAPTIVE": "1",
    "BOT_AUTO_RETRY": "1",
    "BOT_SETTLE_MS": os.environ.get("BOT_SETTLE_MS", "5000"),
    "BOT_BETWEEN_MS": os.environ.get("BOT_BETWEEN_MS", "5000"),
    "BOT_MAX_BETWEEN_MS": os.environ.get("BOT_MAX_BETWEEN_MS", "30000"),
    "BOT_COOLDOWN_FAILS": os.environ.get("BOT_COOLDOWN_FAILS", "5"),
    "BOT_COOLDOWN_MS": os.environ.get("BOT_COOLDOWN_MS", "900000"),   # 15 min throttle cooldown
    "BOT_MAX_PASSES": os.environ.get("BOT_MAX_PASSES", "30"),
    "BOT_DRY_STOP": os.environ.get("BOT_DRY_STOP", "2"),
}
# Rest between full rounds over all files (default 10 min) to let the portal cool.
ROUND_PAUSE_S = float(os.environ.get("BOT_ROUND_PAUSE_MS", "600000")) / 1000.0

_OTP = re.compile(r"\d{3,8}")


def _self_contained_files():
    files = []
    for p in sorted(pathlib.Path(rb.INPUT_DIR).glob("*.xlsx"), key=lambda x: x.stat().st_mtime):
        if p.name.startswith("~$"):
            continue
        cm = rb._colmap_for(p)
        if cm:
            files.append((p, cm))
    return files


def _remaining(path, colmap):
    """(remaining, total) accounts for this file within its configured range."""
    lo, hi = RANGES.get(path.name, (0, 0))
    lo = lo or 2
    done = set()
    res = pathlib.Path(rb.OUTPUT_DIR) / f"OTP_results_{path.stem}.xlsx"
    if res.exists():
        ws = load_workbook(str(res), data_only=True).active
        for r in range(2, ws.max_row + 1):
            cid = rb._txt(ws.cell(r, 1).value)
            if cid and _OTP.fullmatch(rb._txt(ws.cell(r, 2).value)):
                done.add(cid)
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    idc = colmap["id"]
    remaining = total = 0
    for r in range(2, ws.max_row + 1):
        if hi and r > hi:
            break
        if r < lo:
            continue
        cid = rb._txt(ws.cell(r, idc).value)
        if not cid:
            continue
        total += 1
        if cid not in done:
            remaining += 1
    wb.close()
    return remaining, total


def _run_file(path):
    lo, hi = RANGES.get(path.name, (0, 0))
    env = dict(os.environ)
    env.update(OVERNIGHT_ENV)
    env["BOT_FILE"] = path.name
    env["BOT_START_ROW"] = str(lo)
    env["BOT_END_ROW"] = str(hi)
    print(f"[overnight] >>> {path.name}  (rows {lo or 2}..{hi or 'end'})", flush=True)
    subprocess.run([sys.executable, "run_bot.py"], env=env)


def main() -> int:
    print("[overnight] started. Ctrl+C to stop; re-run to resume.", flush=True)
    round_no = 0
    while True:
        round_no += 1
        files = _self_contained_files()
        if not files:
            print("[overnight] no self-contained files in input/ — waiting...", flush=True)
            time.sleep(ROUND_PAUSE_S)
            continue
        worked = False
        for path, cm in files:
            remaining, total = _remaining(path, cm)
            if remaining > 0:
                worked = True
                print(f"[overnight] round {round_no}: {path.name} -> {remaining}/{total} left", flush=True)
                _run_file(path)
            else:
                print(f"[overnight] {path.name}: all {total} captured — skipping", flush=True)
        if not worked:
            print("[overnight] ALL files complete. Nothing left. Exiting.", flush=True)
            return 0
        print(f"[overnight] round {round_no} done; resting {ROUND_PAUSE_S/60:.0f} min before next round...",
              flush=True)
        time.sleep(ROUND_PAUSE_S)


if __name__ == "__main__":
    sys.exit(main())
