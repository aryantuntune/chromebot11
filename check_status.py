from openpyxl import load_workbook
import re, json, pathlib

wb = load_workbook("input/asu-218-0806.xlsx", read_only=True, data_only=True)
target_ids = [str(r[0]).strip() for r in wb.active.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
wb.close()
target_set = set(target_ids)

results = {}
p = pathlib.Path("output/OTP_results.xlsx")
if p.exists():
    wb2 = load_workbook(str(p), read_only=True, data_only=True)
    for row in wb2.active.iter_rows(min_row=2, values_only=True):
        cid = str(row[0] or "").strip()
        val = str(row[1] or "").strip()
        if cid in target_set:
            results[cid] = val
    wb2.close()

RETRY = ("NO_BOOKING_TABLE","NO_BOOKINGS","EMPTY_STATUS","CAPTURE_ERROR","ERROR","CAPTCHA_FAILED")
captured  = [(c,v) for c,v in results.items() if re.fullmatch(r"\d{3,8}", v)]
retryable = [(c,v) for c,v in results.items() if any(v.upper().startswith(m) for m in RETRY)]
other     = [(c,v) for c,v in results.items() if not re.fullmatch(r"\d{3,8}", v) and not any(v.upper().startswith(m) for m in RETRY)]
missing   = [c for c in target_ids if c not in results]

print("Total targets :", len(target_ids))
print("Captured OTPs :", len(captured))
print("Retryable     :", len(retryable))
print("Other (done)  :", len(other))
print("Not yet run   :", len(missing))

sp = pathlib.Path("output/supervisor_status.json")
if sp.exists():
    sup = json.loads(sp.read_text())
    print("Supervisor    : run #" + str(sup.get("run","?")) + "  state=" + str(sup.get("state","?")) + "  stalls=" + str(sup.get("stall_streak",0)))
else:
    print("Supervisor    : no status file yet")

print()
print("--- Captured so far ---")
for c,v in sorted(captured):
    print(" ", c, "->", v)

if retryable:
    print()
    print("--- Still retryable ---")
    for c,v in sorted(retryable):
        print(" ", c, "->", v[:60])
