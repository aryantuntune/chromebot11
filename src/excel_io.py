"""Excel input/output helpers for the sign-in / scrape bot.

This module is responsible for everything that touches the spreadsheet:

* discovering the input workbook,
* loading the rows of credentials,
* writing scraped results back, and
* saving a *copy* of the workbook into the output directory.

It uses ``openpyxl`` exclusively and never modifies the original input file.
"""

from __future__ import annotations

import pathlib
from dataclasses import dataclass

from openpyxl import load_workbook as _openpyxl_load_workbook
from openpyxl.utils import column_index_from_string

# Imported only for the type hint on ``excel_cfg`` parameters.  Importing the
# real dataclass keeps the contract honest: callers pass the shared
# ``ExcelConfig`` from ``src.config``.
from .config import ExcelConfig


@dataclass
class RowData:
    """A single data row read from the worksheet."""

    row_number: int          # 1-based worksheet row
    email: str
    password: str
    result: str | None


def find_input_file(input_dir: str) -> pathlib.Path:
    """Return the most recently modified ``.xlsx`` file in ``input_dir``.

    Excel temporary/lock files (whose names start with ``~$``) are ignored.

    Raises:
        FileNotFoundError: if no suitable ``.xlsx`` file exists.
    """
    directory = pathlib.Path(input_dir)
    if not directory.is_dir():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir!r}")

    candidates = [
        p
        for p in directory.glob("*.xlsx")
        if p.is_file() and not p.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No .xlsx input files found in directory: {input_dir!r}"
        )

    # Most recently modified wins.
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _cell_text(value: object) -> str:
    """Coerce a cell value to a stripped string (``None`` -> empty string)."""
    if value is None:
        return ""
    return str(value).strip()


def load_workbook_and_rows(path, excel_cfg: ExcelConfig):
    """Open the workbook at ``path`` and read its credential rows.

    Args:
        path: path to the ``.xlsx`` file (``str`` or ``pathlib.Path``).
        excel_cfg: the shared :class:`~src.config.ExcelConfig`.

    Returns:
        A tuple ``(workbook, worksheet, rows)`` where ``rows`` is a list of
        :class:`RowData`.  Rows are read from ``first_data_row`` through
        ``worksheet.max_row`` (inclusive); rows whose email cell is empty are
        skipped.

    The workbook is opened with ``data_only=False`` so that it can be written
    back and saved later without losing formulas.
    """
    # ``data_only=False`` keeps the workbook writable for save-back.
    workbook = _openpyxl_load_workbook(filename=str(path), data_only=False)

    if excel_cfg.sheet_name is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[excel_cfg.sheet_name]

    email_idx = column_index_from_string(excel_cfg.email_col)
    password_idx = column_index_from_string(excel_cfg.password_col)
    result_idx = column_index_from_string(excel_cfg.result_col)

    rows: list[RowData] = []
    for row_number in range(excel_cfg.first_data_row, worksheet.max_row + 1):
        email = _cell_text(worksheet.cell(row=row_number, column=email_idx).value)
        # Skip rows with no email -- these are blank/padding rows.
        if not email:
            continue

        password = _cell_text(
            worksheet.cell(row=row_number, column=password_idx).value
        )
        result_raw = worksheet.cell(row=row_number, column=result_idx).value
        result = None if result_raw is None else str(result_raw)

        rows.append(
            RowData(
                row_number=row_number,
                email=email,
                password=password,
                result=result,
            )
        )

    return workbook, worksheet, rows


def write_result(ws, excel_cfg: ExcelConfig, row_number: int, value: str) -> None:
    """Write ``value`` into ``result_col`` at ``row_number`` on ``ws``."""
    result_idx = column_index_from_string(excel_cfg.result_col)
    ws.cell(row=row_number, column=result_idx, value=value)


def save_workbook(wb, input_path, output_dir: str) -> pathlib.Path:
    """Save a copy of ``wb`` into ``output_dir`` and return the output path.

    The output filename is derived deterministically from the input filename
    (``"<stem>.result.xlsx"``) so that repeated/incremental saves during a run
    overwrite the *same* file rather than producing many files.

    The original input file is never touched.
    """
    out_dir = pathlib.Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    stem = pathlib.Path(input_path).stem
    output_path = out_dir / f"{stem}.result.xlsx"
    wb.save(str(output_path))
    return output_path
